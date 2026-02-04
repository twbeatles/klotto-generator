"""
DB 데이터를 엑셀 파일로 내보내는 스크립트
"""
import sqlite3
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 패키지가 필요합니다. 설치 중...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

# Configuration
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "lotto_history.db"

def export_to_excel(output_path: Path = None):
    """Export database to Excel file."""
    if not DB_PATH.exists():
        print(f"데이터베이스를 찾을 수 없습니다: {DB_PATH}")
        return False

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = DATA_DIR / f"lotto_history_{timestamp}.xlsx"

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Fetch all data
        cursor.execute("""
            SELECT draw_no, date, num1, num2, num3, num4, num5, num6, bonus, 
                   prize_amount, winners_count, total_sales 
            FROM draws 
            ORDER BY draw_no ASC
        """)
        rows = cursor.fetchall()
        conn.close()

        if not rows:
            print("내보낼 데이터가 없습니다.")
            return False

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "로또 당첨번호"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "회차", "추첨일", "번호1", "번호2", "번호3", "번호4", "번호5", "번호6", 
            "보너스", "1등 상금(원)", "1등 당첨자", "총 판매액(원)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # Format currency columns
                if col_idx in [10, 12] and value:  # prize_amount, total_sales
                    cell.number_format = '#,##0'

        # Adjust column widths
        column_widths = [8, 12, 8, 8, 8, 8, 8, 8, 8, 18, 12, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save
        wb.save(output_path)
        print(f"엑셀 파일 저장 완료: {output_path}")
        print(f"총 {len(rows)}개 회차 데이터 내보내기 완료")
        return True

    except Exception as e:
        print(f"엑셀 내보내기 실패: {e}")
        return False

if __name__ == "__main__":
    export_to_excel()
